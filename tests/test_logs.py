"""
Tests for converted operator logs — python tests/test_logs.py

Plain asserts to match the calculator tests; this repo has no pytest.

These read the manifest and parquet produced by logs/convert.py. They do not
call a model and do not convert anything, so they are fast and deterministic —
but they only pass once `python -m logs.convert` has been run.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from logs import catalogue, find, load_log, read_manifest  # noqa: E402
from logs.schema import validate_frame  # noqa: E402

WB = (
    Path(__file__).resolve().parent.parent
    / "documents" / "test"
    / "DWQMS-Drinking-Water-System-Record-Template.xlsx"
)

entries = read_manifest()
if not entries:
    print("test_logs: SKIPPED - no manifest. Run: python -m logs.convert --input "
          "documents/test/DWQMS-Drinking-Water-System-Record-Template.xlsx")
    raise SystemExit(0)

converted = {e.sheet_name: e for e in entries if e.status == "converted"}

# 1. The twelve monthly tabs converted, with the real length of each month —
#    including February of a leap year. That is what proves the converter
#    detects where data ends rather than skipping a fixed number of footer rows.
MONTHS = {"Jan.": 31, "Feb.": 29, "March": 31, "April": 30, "May": 31, "June": 30,
          "July": 31, "Aug.": 31, "Sept.": 30, "Oct.": 31, "Nov.": 30, "Dec.": 31}
for sheet, expected in MONTHS.items():
    assert sheet in converted, f"{sheet} not converted"
    assert converted[sheet].row_count == expected, (sheet, converted[sheet].row_count)

# 2. Reading a log touches parquet only — no model call, no exec, no openpyxl.
jan = load_log("Jan.")
# Rows are pinned because January has 31 days whatever the converter does.
# Columns are not: the converter is rewritten by the model on every conversion,
# and how it joins a multi-row header ("Flows | Raw Water | Meter Reading" as
# one column or three) legitimately shifts the count - it has been 127 and 133
# across two runs of the same workbook. Pinning it makes re-converting the
# corpus fail a test that is not measuring anything about correctness.
assert isinstance(jan, pd.DataFrame), type(jan)
assert jan.shape[0] == 31, jan.shape
assert jan.shape[1] > 50, jan.shape
assert isinstance(jan.index, pd.DatetimeIndex) and jan.index.name == "date"

# 3. Column names are unique and carry the level above them. Six columns on this
#    sheet are labelled "Meter Reading (m3)" at the leaf and are distinguishable
#    only by their group and block.
assert len(set(jan.columns)) == len(jan.columns)
meter = [c for c in jan.columns if c.endswith("Meter Reading (m3)")]
assert len(meter) == 6, meter
assert any("Raw Water" in c for c in meter) and any("Treated Water" in c for c in meter)

# 4. The date is reconstructed, and a typo in one of the sheet's fourteen
#    redundant Date columns does not corrupt it. Jan. row 8 reads 11 in column A
#    and 1 in the other thirteen; the day the majority agree on is the right one.
assert 1 in set(jan.index.day), "day 1 lost to the column-A typo"
assert not jan.index.has_duplicates
assert jan.index.min().year == 2016 and jan.index.min().month == 1

# 5. Values survive the round trip to parquet unchanged.
ws = load_workbook(WB, data_only=True)["Jan."]
turbidity = next(c for c in jan.columns if "Raw Water" in c and "Turbidity" in c)
assert jan.loc["2016-01-02", turbidity] == ws["BB9"].value == 0.788

# 6. Every converted frame passes the same validation the conversion applied.
for sheet, entry in converted.items():
    failures, _ = validate_frame(load_log(sheet))
    assert not failures, (sheet, failures)

# 7. cell_coverage records how much of the source landed in the parquet. It is
#    reported, never enforced: Nov. and Dec. score 0.0 because those sheets are
#    blank templates carrying only pre-filled day numbers, which is the truth
#    about the workbook rather than a conversion fault.
assert converted["Sept."].cell_coverage > 0.7, converted["Sept."].cell_coverage
assert converted["Nov."].cell_coverage == 0.0
assert load_log("Nov.").notna().sum().sum() == 0

# 8. Every entry can be traced back to the sheet it came from, and to the
#    converter that produced it.
for entry in entries:
    assert entry.file_path and entry.file_hash and entry.sheet_name
    if entry.status == "converted":
        assert Path(entry.parquet_path).exists(), entry.parquet_path
        assert "def convert" in entry.converter_source

# 9. The catalogue is what goes into context, so it must name every converted
#    sheet and carry a description for each.
cat = catalogue()
for sheet, entry in converted.items():
    assert f"sheet: {sheet}" in cat, sheet
    assert entry.description, f"{sheet} has no description"

# 10. A sheet the converter could not handle is recorded, not silently dropped.
unconverted = [e for e in entries if e.status != "converted"]
for entry in unconverted:
    assert entry.error, entry.sheet_name
    assert find(entry.sheet_name) is None or find(entry.sheet_name).status == "converted"

print(f"test_logs: all assertions passed "
      f"({len(converted)} converted, {len(unconverted)} unconverted)")
