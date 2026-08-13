#!/usr/bin/env python3
"""Build the retrieval index: documents and interpreted spreadsheets -> Chroma.

Two configs are supported and they differ in EXACTLY ONE THING: the table
serializer docling uses while chunking. Everything else - the corpus, the
chunker, its token budget, the embedding model, the schema - is identical, so a
difference in retrieval scores between them is attributable to that one choice
and nothing else. There is no config object and no registry: a flag and a
branch, so the claim "only the serializer varies" is checkable by reading
twenty lines rather than trusted.

    python scripts/build_index.py --input documents/rag_test \\
        --config-id triplet_tables --plant-id demo --table-serializer triplet

    python scripts/build_index.py --input documents/rag_test \\
        --config-id markdown_tables --plant-id demo --table-serializer markdown

--dry-run writes the chunks it would embed and calls nothing - no embeddings
API, no database. Run it first: if the chunk text is wrong, no amount of
retrieval tuning downstream will fix it, and you will have paid to find out.

Two input paths, because two kinds of source need different treatment:

  PDFs and .docx go through docling. If inspect_parse.py has already written a
  DoclingDocument for the file, that JSON is reused rather than reconverted -
  a 57-page scan costs minutes to convert and the parse is the thing you
  already reviewed. Anything without a cached parse is converted here using
  inspect_parse's own converter settings, imported rather than copied so the
  two cannot drift.

  Spreadsheets do NOT go through docling - it flattens them into a bare grid
  and drops the formulas and merged ranges that carry the meaning. They come
  from the reviewed output of interpret_sheets.py instead, where each `durable`
  statement is already a self-contained sentence and therefore already a chunk.
  `instance_values` are ignored entirely: they are one run of a sheet, not
  knowledge, and indexing today's inlet temperature as a plant fact is how a
  retrieval system starts confidently reporting last Tuesday.
"""

import argparse
import json
import sys
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Imported before anything else touches torch: inspect_parse sets the dynamo
# environment variables at import time, and they are only read once.
import inspect_parse  # noqa: E402

from docling_core.transforms.chunker.hierarchical_chunker import (  # noqa: E402
    ChunkingDocSerializer,
    ChunkingSerializerProvider,
    TripletTableSerializer,
)
from docling_core.transforms.chunker.hybrid_chunker import HybridChunker  # noqa: E402
from docling_core.transforms.chunker.tokenizer.huggingface import (  # noqa: E402
    HuggingFaceTokenizer,
)
from docling_core.transforms.serializer.common import create_ser_result  # noqa: E402
from docling_core.transforms.serializer.markdown import (  # noqa: E402
    MarkdownTableSerializer,
)
from docling_core.types.doc import DoclingDocument  # noqa: E402

from rag import clean, embed, store  # noqa: E402

DOC_FORMATS = {".pdf", ".docx"}
SHEET_FORMATS = {".xlsx", ".xlsm"}

# Counts tokens to size chunks. NOT the embedding model - it never sees a
# vector. The two tokenizers disagree, so a chunk sized at N here is only
# approximately N to the embedding API; that is harmless because 512 of these
# is nowhere near text-embedding-3's 8191-token input limit, so nothing is ever
# silently truncated on the way in.
CHUNK_TOKENIZER = "sentence-transformers/all-MiniLM-L6-v2"

# docling's own default is 256, which is too small for this corpus: these SOPs
# carry long heading trails, and at 256 the contextualizing prefix alone
# overflows the budget and gets split across chunks, so later chunks lose the
# heading that says which procedure they belong to. 512 is the tokenizer's own
# maximum and fits the prefix. Identical for both configs, so it cannot explain
# a difference between them.
DEFAULT_MAX_TOKENS = 512


# --------------------------------------------------------------------------
# The one variable
# --------------------------------------------------------------------------
# A table has to become a string before it can be embedded, and there is no
# neutral way to do it. Triplets state each cell with its row and column
# headers, so every fact survives on its own and the chunk reads as prose;
# markdown keeps the grid, so spatial relationships survive but a retrieved
# fragment may be a row whose header row got chunked away. Which one retrieves
# better on operations documents is an empirical question, which is why both
# exist here instead of one being chosen in advance.



# Both serializers normalise their own output BEFORE returning it, which is the
# only way the comparison can be fair. The chunker sizes chunks by tokenizing
# what the serializer emits, and docling pads markdown tables out to the widest
# cell in every column. Measured on this corpus, that padding alone gave
# markdown 2.1x the table chunks at 59% the size - the chunk boundaries were
# being set by whitespace, not content, and markdown would have lost a
# comparison it never actually ran. Normalising here means both configs are
# measured on what their representation says, not on how much it indents.
#
# The post-chunk normalise() in chunks_from_documents still runs and is still
# needed: it cleans the prose the chunker merges around these tables. It is
# idempotent, so text passing through both paths is unharmed.


class NormalisedTripletTableSerializer(TripletTableSerializer):
    def serialize(self, *, item, doc_serializer, doc, **kwargs):
        result = super().serialize(
            item=item, doc_serializer=doc_serializer, doc=doc, **kwargs
        )
        return create_ser_result(text=clean.normalise(result.text), span_source=item)


class NormalisedMarkdownTableSerializer(MarkdownTableSerializer):
    def serialize(self, *, item, doc_serializer, doc, **kwargs):
        result = super().serialize(
            item=item, doc_serializer=doc_serializer, doc=doc, **kwargs
        )
        return create_ser_result(text=clean.normalise(result.text), span_source=item)


class TripletTableProvider(ChunkingSerializerProvider):
    def get_serializer(self, doc):
        return ChunkingDocSerializer(
            doc=doc, table_serializer=NormalisedTripletTableSerializer()
        )


class MarkdownTableProvider(ChunkingSerializerProvider):
    def get_serializer(self, doc):
        return ChunkingDocSerializer(
            doc=doc, table_serializer=NormalisedMarkdownTableSerializer()
        )


def serializer_provider(name):
    return TripletTableProvider() if name == "triplet" else MarkdownTableProvider()


# --------------------------------------------------------------------------
# Documents
# --------------------------------------------------------------------------


def load_document(path, parsed_dir, converter):
    """Return (DoclingDocument, source) where source is 'cached' or 'converted'."""
    cached = parsed_dir / f"{path.stem}.json"
    if cached.is_file():
        return (
            DoclingDocument.model_validate_json(cached.read_text(encoding="utf-8")),
            "cached",
        )
    return converter().convert(path).document, "converted"


def page_span(chunk):
    """'p.7', or 'pp.7-8' when a chunk straddles a page break.

    Empty for formats with no pagination (.docx), which is why the caller has a
    fallback rather than writing an empty location into a NOT NULL column.
    """
    pages = sorted(
        {
            prov.page_no
            for item in chunk.meta.doc_items
            for prov in (getattr(item, "prov", None) or [])
            if getattr(prov, "page_no", None) is not None
        }
    )
    if not pages:
        return ""
    if len(pages) == 1:
        return f"p.{pages[0]}"
    return f"pp.{pages[0]}-{pages[-1]}"


def chunk_content_type(chunk):
    labels = {
        getattr(item.label, "value", str(item.label)) for item in chunk.meta.doc_items
    }
    return "table" if "table" in labels else "prose"


def chunks_from_documents(paths, parsed_dir, provider, max_tokens, log):
    """Chunk every document with the HybridChunker under the chosen serializer."""
    chunker = HybridChunker(
        serializer_provider=provider,
        tokenizer=HuggingFaceTokenizer.from_pretrained(
            CHUNK_TOKENIZER, max_tokens=max_tokens
        ),
    )

    # Built on first use only. A fully cached corpus should not pay for
    # instantiating the layout and table models it will never call.
    cache = {}

    def converter():
        if "converter" not in cache:
            cache["converter"] = inspect_parse.build_converter(
                SimpleNamespace(table_mode="accurate", no_cell_matching=False, no_ocr=False)
            )
        return cache["converter"]

    rows = []
    for path in paths:
        try:
            doc, source = load_document(path, parsed_dir, converter)
        except Exception as exc:
            log(f"  {path.name}: FAILED to load: {type(exc).__name__}: {exc}")
            continue

        made = 0
        for ordinal, chunk in enumerate(chunker.chunk(dl_doc=doc)):
            headings = list(getattr(chunk.meta, "headings", None) or [])
            location = page_span(chunk)
            if not location:
                location = f"section '{headings[-1]}'" if headings else f"chunk {ordinal}"
            rows.append(
                {
                    "chunk_id": f"{path.name}#{ordinal:04d}",
                    "document": path.name,
                    "location": location,
                    "section": " > ".join(headings),
                    "content_type": chunk_content_type(chunk),
                    # contextualize() prepends the heading trail. The retrieved
                    # text has to stand alone - a chunk reading "0.2 mg/L" with
                    # its section stripped is unusable no matter how well it
                    # scored - so this, not chunk.text, is what gets embedded.
                    "text": clean.normalise(chunker.contextualize(chunk=chunk)),
                }
            )
            made += 1
        log(f"  {path.name}: {made} chunks ({source})")
    return rows


# --------------------------------------------------------------------------
# Spreadsheets
# --------------------------------------------------------------------------


def sanitise(name):
    """interpret_sheets.py's filename transform, character for character.

    Copied deliberately rather than imported: interpret_sheets.py is not to be
    modified, and importing it would pull in its module-level openpyxl and
    argparse setup for the sake of one string transform. If that naming ever
    changes, this is the line that has to change with it.
    """
    return name.replace(" ", "_").replace("/", "_")


def resolve_sheet(workbooks, stem):
    """Map an interp filename stem back to (workbook path, real sheet name).

    The sanitised stem is lossy - spaces became underscores - so the sheet name
    is recovered by re-sanitising the workbook's actual sheet names and looking
    for the one that matches, rather than by guessing the inverse transform.
    The transform is per-character, so sanitise(stem + "__" + sheet) always
    splits as sanitise(stem + "__") + sanitise(sheet).
    """
    from openpyxl import load_workbook

    for path in workbooks:
        prefix = sanitise(f"{path.stem}__")
        if not stem.startswith(prefix):
            continue
        key = stem[len(prefix) :]
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in book.sheetnames:
                if sanitise(sheet_name) == key:
                    return path, sheet_name
        finally:
            book.close()
    return None, None


def statement_location(sheet_name, source_cells):
    cells = [c for c in (source_cells or []) if c]
    if not cells:
        return f"sheet '{sheet_name}'"
    return f"sheet '{sheet_name}', cells {', '.join(cells)}"


def chunks_from_sheets(workbooks, interp_dir, log):
    """One chunk per reviewed `durable` statement. instance_values are dropped."""
    if not interp_dir.is_dir():
        log(f"  no interpretation directory at {interp_dir} - no sheet chunks")
        return []

    rows = []
    counters = {}
    unmatched = 0
    for json_path in sorted(interp_dir.glob("*.json")):
        path, sheet_name = resolve_sheet(workbooks, json_path.stem)
        if path is None:
            unmatched += 1
            continue

        result = json.loads(json_path.read_text(encoding="utf-8"))
        durable = result.get("durable", [])
        dropped = len(result.get("instance_values", []))
        unsure = [d for d in durable if d.get("confidence") != "high"]

        for item in durable:
            ordinal = counters.get(path.name, 0)
            counters[path.name] = ordinal + 1
            rows.append(
                {
                    "chunk_id": f"{path.name}#{ordinal:04d}",
                    "document": path.name,
                    "location": statement_location(sheet_name, item.get("source_cells")),
                    "section": item.get("category", ""),
                    "content_type": "interpreted_statement",
                    # Already clean prose - these were written as sentences,
                    # not extracted from a page - but normalised anyway so
                    # every chunk in the index went through one path.
                    "text": clean.normalise(item["statement"]),
                }
            )
        note = f" - {len(unsure)} NOT high confidence, check the .review.md" if unsure else ""
        log(
            f"  {path.name} [{sheet_name}]: {len(durable)} statements, "
            f"{dropped} instance values ignored{note}"
        )

    if unmatched:
        log(
            f"  {unmatched} interpretation file(s) in {interp_dir} belong to "
            "workbooks outside --input and were skipped"
        )
    return rows


# --------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Chunk a corpus and write it to Postgres/pgvector."
    )
    parser.add_argument("--input", type=Path, required=True, help="corpus directory")
    parser.add_argument("--config-id", required=True, help="names this ingestion config")
    parser.add_argument("--plant-id", required=True)
    parser.add_argument(
        "--table-serializer",
        choices=("triplet", "markdown"),
        required=True,
        help="the ONLY thing that differs between the two configs",
    )
    parser.add_argument(
        "--parsed-dir",
        type=Path,
        default=ROOT / "data" / "parsed",
        help="DoclingDocument JSON from inspect_parse.py (reused when present)",
    )
    parser.add_argument(
        "--interp-dir",
        type=Path,
        default=ROOT / "data" / "interp",
        help="reviewed output of interpret_sheets.py",
    )
    parser.add_argument("--embedding-model", default=embed.DEFAULT_MODEL)
    parser.add_argument(
        "--store-path",
        type=Path,
        default=None,
        help=f"Chroma directory (default {store.DEFAULT_PATH}); both configs "
        "must share one, since the comparison is a filter within it",
    )
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=DEFAULT_MAX_TOKENS,
        help=f"chunk token budget (default {DEFAULT_MAX_TOKENS}); "
        "must be the same for both configs or the comparison is meaningless",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="write the chunks that would be embedded and call nothing",
    )
    parser.add_argument(
        "--preview-dir",
        type=Path,
        default=ROOT / "data" / "chunks",
        help="where --dry-run writes its .jsonl",
    )
    args = parser.parse_args()

    if not args.input.is_dir():
        raise SystemExit(f"input directory not found: {args.input}")

    docs = sorted(p for p in args.input.iterdir() if p.suffix.lower() in DOC_FORMATS)
    sheets = sorted(p for p in args.input.iterdir() if p.suffix.lower() in SHEET_FORMATS)
    others = sorted(
        p
        for p in args.input.iterdir()
        if p.is_file() and p.suffix.lower() not in DOC_FORMATS | SHEET_FORMATS
    )

    def log(line):
        print(line, flush=True)

    log(f"config {args.config_id!r}, plant {args.plant_id!r}")
    log(f"table serializer: {args.table_serializer}")
    log(f"{len(docs)} document(s), {len(sheets)} spreadsheet(s) in {args.input}")
    if others:
        # Named rather than ignored: .xls in particular looks handled and is
        # not - openpyxl cannot read it, so interpret_sheets.py never produced
        # an interpretation for it and there is nothing here to index.
        log(
            f"skipping {len(others)} unsupported file(s): "
            + ", ".join(p.name for p in others[:5])
            + (" ..." if len(others) > 5 else "")
        )

    log("\ndocuments:")
    rows = chunks_from_documents(
        docs,
        args.parsed_dir,
        serializer_provider(args.table_serializer),
        args.max_tokens,
        log,
    )
    log("\nspreadsheets:")
    rows += chunks_from_sheets(sheets, args.interp_dir, log)

    # A chunk that was nothing but a horizontal rule normalises to empty. The
    # embeddings API rejects empty input, so these are dropped here and
    # counted - dropping them silently would make the chunk totals in this
    # report disagree with what is actually in the index.
    empty = [r for r in rows if not r["text"]]
    if empty:
        rows = [r for r in rows if r["text"]]
        log(f"\ndropped {len(empty)} chunk(s) left empty by normalisation: "
            + ", ".join(r["chunk_id"] for r in empty[:5])
            + (" ..." if len(empty) > 5 else ""))

    if not rows:
        raise SystemExit("\nno chunks produced - nothing to write")

    for row in rows:
        row["config_id"] = args.config_id
        row["plant_id"] = args.plant_id

    by_type = {}
    for row in rows:
        by_type[row["content_type"]] = by_type.get(row["content_type"], 0) + 1
    log(f"\n{len(rows)} chunks: " + ", ".join(f"{k} {v}" for k, v in sorted(by_type.items())))

    if args.dry_run:
        args.preview_dir.mkdir(parents=True, exist_ok=True)
        out = args.preview_dir / f"{args.config_id}__{args.plant_id}.jsonl"
        with out.open("w", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
        log("\ndry run - nothing embedded, nothing written to the database")
        log(f"-> {out}   read this before paying to embed it")
        return 0

    dim = embed.dimension(args.embedding_model)
    log(f"\nembedding {len(rows)} chunks with {args.embedding_model} ({dim}d)")
    vectors = embed.embed_texts(
        [r["text"] for r in rows],
        model=args.embedding_model,
        progress=lambda done, total: print(f"  {done}/{total}", flush=True),
    )
    for row, vector in zip(rows, vectors):
        row["embedding"] = vector
        row["embedding_model"] = args.embedding_model

    _client, collection = store.open_collection(args.store_path)
    store.check_embedding_model(collection, args.embedding_model)
    written, deleted = store.write_chunks(collection, rows, args.config_id, args.plant_id)
    log(f"\nwrote {written} chunks to {args.store_path or store.default_path()}, "
        f"removed {deleted} stale")
    for content_type, count in store.scope_summary(collection, args.config_id, args.plant_id):
        log(f"  {content_type}: {count}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
