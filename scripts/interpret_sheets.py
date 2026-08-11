#!/usr/bin/env python3
"""Spreadsheet interpretation - openpyxl extracts, Claude interprets, you approve.

Runs after inspect_sheets.py. openpyxl supplies everything mechanical (cell
coordinates, values, formulas, merged ranges, fill colours); the model supplies
the one thing no rule recovers: which label three rows up and one column left
belongs to which value, and what a formula therefore MEANS.

Output is a DRAFT for review, not a finished artefact. Every claim carries the
cells it was drawn from and a confidence, so checking it is spot-checking a
citation rather than re-deriving the sheet. Read the .review.md, fix what is
wrong, and only then let the result feed chunk generation.

    export ANTHROPIC_API_KEY=...            # or setx on Windows
    python scripts/interpret_sheets.py --input data/corpus --output data/interp
    python scripts/interpret_sheets.py --input data/corpus --dry-run

--dry-run writes the exact payload that would be sent and calls nothing. Use it
first: if the extracted cell dump is missing formulas or labels, the model
cannot recover them and no amount of prompting will help.

The split that matters: DURABLE knowledge (equations, variable definitions,
fixed plant geometry, thresholds, decision rules) is what belongs in a
knowledge base. INSTANCE values (today's temperature in an input cell) are one
run of the sheet and are captured separately so they can be excluded.

Sheets that are one form filled in over and over - a month-per-tab record
workbook - are handled as a group: the first is interpreted and the rest are
skipped by name on stdout. Twelve monthly tabs share one template, so the
durable content (licence limits, tank geometry, chemical strengths, all of it
living in the formulas rather than the entries) is captured once instead of
twelve times. Pass --include-record-grids to interpret every one of them.
"""

import argparse
import json
import os
import re
import sys
import traceback
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
except ImportError:  # optional - the shell environment works just as well
    def load_dotenv(*_args, **_kwargs):
        return False

# Counts CELLS, not rows. A wide matrix hits a low cap at a fraction of the
# row count you would expect: 22 columns x 50 rows is already ~1100 cells even
# when sparsely filled. Truncation is silent to the model - it interprets the
# fragment it was given and has no way to know the rest existed - so this is
# set high enough to pass a whole normal sheet and the run says so loudly when
# it still trips.
MAX_CELLS = 5000

# Per-cell caps. These exist to stop one pathological cell from swallowing the
# payload, NOT to save tokens - a header cell that lists every critical control
# limit on the sheet is exactly the durable knowledge this script is for, and
# cutting it mid-sentence loses thresholds silently. The longest value in the
# current corpus is 201 chars, so 1000 costs nothing and any cell that still
# trips it is reported by coordinate.
MAX_VALUE_CHARS = 1000
MAX_FORMULA_CHARS = 1000

# A record grid is many copies of one formula: the monthly log sheets carry
# ~1571 formula cells built from 25 distinct shapes, 440 of them =IF(A1="","",A1).
# A real calculator is the opposite - the CT sheet has 8 formulas in 7 shapes.
# Both numbers are needed: the ratio alone would flag a small calculator that
# happens to repeat one row, and the count alone would flag any large sheet.
MIN_FORMULAS_FOR_GRID = 50
MAX_DISTINCT_SHAPE_RATIO = 0.10

# Two record grids belong to the same group when their formula shapes largely
# agree. Not exact equality: each monthly tab reaches into the NEXT month for
# the opening meter reading, so Jan. carries a Feb. reference that Feb. does
# not, and December - with no January after it - is a 22-shape subset of the
# other months' 25. Jaccard 0.8 spans that without merging genuinely different
# sheets (Year-to-Date and Sampling Results score far below it).
GROUP_SIMILARITY = 0.8

# Ceiling on the interpretation itself. The richest sheet so far produces about
# 2800 tokens, so there is room - but a sheet whose every row deserves its own
# statement (the 39x30 hazard matrix, if it is ever chunked row-wise) would
# reach this, and a response cut here cannot be parsed or partially recovered.
MAX_OUTPUT_TOKENS = 8000

SHEET_FORMATS = {".xlsx", ".xlsm"}

SCHEMA = {
    "type": "object",
    "properties": {
        "sheet_purpose": {
            "type": "string",
            "description": "One sentence: what this sheet is for, in plant terms.",
        },
        "kind": {
            "type": "string",
            "enum": ["calculator", "reference", "log", "mixed", "unclear"],
        },
        "durable": {
            "type": "array",
            "description": "Knowledge that stays true after the input cells change.",
            "items": {
                "type": "object",
                "properties": {
                    "statement": {
                        "type": "string",
                        "description": (
                            "A self-contained sentence readable on its own, with "
                            "no reference to cell coordinates or 'this sheet'."
                        ),
                    },
                    "category": {
                        "type": "string",
                        "enum": [
                            "equation",
                            "definition",
                            "constant",
                            "plant_geometry",
                            "threshold",
                            "decision_rule",
                        ],
                    },
                    "formula": {
                        "type": "string",
                        "description": "Verbatim formula if one underlies this, else empty.",
                    },
                    "source_cells": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Cells this was drawn from, e.g. ['A8','C8'].",
                    },
                    "evidence": {
                        "type": "string",
                        "description": (
                            "Which cell supplied the LABEL and which supplied the "
                            "VALUE or formula. This is what makes review fast."
                        ),
                    },
                    "confidence": {
                        "type": "string",
                        "enum": ["high", "medium", "low"],
                    },
                },
                "required": [
                    "statement",
                    "category",
                    "formula",
                    "source_cells",
                    "evidence",
                    "confidence",
                ],
                "additionalProperties": False,
            },
        },
        "instance_values": {
            "type": "array",
            "description": "Entered or calculated values for one run. NOT knowledge.",
            "items": {
                "type": "object",
                "properties": {
                    "label": {"type": "string"},
                    "cell": {"type": "string"},
                    "value": {"type": "string"},
                },
                "required": ["label", "cell", "value"],
                "additionalProperties": False,
            },
        },
        "uncertain": {
            "type": "array",
            "description": "Cells you could not confidently interpret. Do not guess.",
            "items": {
                "type": "object",
                "properties": {
                    "cells": {"type": "array", "items": {"type": "string"}},
                    "question": {"type": "string"},
                },
                "required": ["cells", "question"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["sheet_purpose", "kind", "durable", "instance_values", "uncertain"],
    "additionalProperties": False,
}

SYSTEM = """You interpret spreadsheets from drinking-water treatment plant \
documentation so their knowledge can be retrieved later by plant operators.

You are given a cell-by-cell dump: coordinates, displayed values, underlying \
formulas, merged ranges and fill colours. Position encodes meaning - a label \
often sits several rows above or a column left of the value it names, and a \
merged range usually spans the block it heads. Use the formula's cell \
references to work out what a calculated cell depends on.

Your central task is separating DURABLE knowledge from ONE INSTANCE of it:

- The equation CT = 0.353*I*(12.006+e^(...)) is durable.
- The 7.5 C someone typed into the input cell today is an instance.
- The clearwell being 155 m2 with a baffling factor of 0.5 is durable - it is \
a physical fact about the plant.
- A cell computing a result from today's inputs is an instance, but the \
FORMULA behind it is durable.

Rules:
- Write each durable statement so it stands alone. An operator retrieving it \
sees only that sentence - no cell references, no "as shown above".
- Every claim cites the cells it came from, and evidence naming which cell \
gave the label and which gave the value.
- Where a fill colour legend exists on the sheet, use it - it usually marks \
entered vs calculated cells.
- When a label-to-value association is ambiguous, put it in `uncertain` with \
the question you would ask. Do NOT guess. A confident wrong statement in an \
operations knowledge base is worse than an acknowledged gap.
- Mark confidence honestly. Reserve `high` for cases where the label sits \
adjacent to the value or the formula makes the meaning unambiguous."""

def fill_colour(cell):
    fill = getattr(cell, "fill", None)
    if fill is None or fill.patternType != "solid":
        return None
    colour = fill.start_color
    if getattr(colour, "type", None) == "rgb" and colour.rgb:
        rgb = str(colour.rgb)
        return None if rgb in ("00000000", "FFFFFFFF") else rgb
    if getattr(colour, "type", None) == "theme":
        return f"theme:{colour.theme}"
    return None


CELL_REF = re.compile(r"\$?[A-Z]{1,3}\$?\d{1,7}")
SHEET_REF = re.compile(r"(?:'[^']+'|[A-Za-z_][\w. ]*)!")


def formula_shape(formula):
    """A formula with its cell and sheet references blanked out.

    =IF(A1="","",A1) and =IF(B7="","",B7) are the same rule filled down a
    column; collapsing the references is what lets them be counted as one.
    Sheet names go too, so that Jan.'s roll-over into Feb. and Feb.'s into
    March read as the same rule rather than two - which is what lets the
    monthly tabs be recognised as one repeated template.
    """
    return CELL_REF.sub("@", SHEET_REF.sub("@!", formula))


def record_grid_shapes(ws_formulas):
    """Distinct formula shapes if this sheet is one form stamped out many
    times, else None.

    Interpreting these cell by cell pays full price for a sheet whose durable
    content is a single template. The triage script cannot be used for this
    call: its classifier tests `formulas >= 5` first, so it labels every one of
    these monthly sheets a CALCULATOR.
    """
    formulas = [
        cell.value
        for row in ws_formulas.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    if len(formulas) < MIN_FORMULAS_FOR_GRID:
        return None
    shapes = Counter(formula_shape(f) for f in formulas)
    if len(shapes) / len(formulas) >= MAX_DISTINCT_SHAPE_RATIO:
        return None
    return shapes


def group_record_grids(sheet_shapes):
    """Map each record-grid sheet to the sheet that stands in for it.

    A sheet maps to itself when it is the first of its kind seen; every later
    sheet whose formula shapes overlap it maps back to that one. Interpreting
    the representative captures the template all of them share - the licence
    limits and tank geometry live in the formulas, so skipping the whole group
    would throw that away to save a cost that only one call ever needed.
    """
    representatives = []  # (name, shape set), in workbook order
    assignment = {}
    for name, shapes in sheet_shapes.items():
        keys = set(shapes)
        for rep_name, rep_keys in representatives:
            union = keys | rep_keys
            if union and len(keys & rep_keys) / len(union) >= GROUP_SIMILARITY:
                assignment[name] = rep_name
                break
        else:
            representatives.append((name, keys))
            assignment[name] = name
    return assignment


def extract_sheet(ws_values, ws_formulas):
    """Flat text dump of one sheet: coordinate, value, formula, fill.

    Deliberately NOT a rendered grid - a wide sparse sheet wastes most of the
    payload on padding. One line per non-empty cell keeps coordinates explicit,
    which is what the model needs to reason about spatial association.

    Returns the payload plus a list of warnings. Anything the extractor drops
    is invisible downstream - the model cannot ask for what it was never
    shown - so every cut is named here rather than made quietly.
    """
    lines = []
    truncated = False
    clipped = []
    for row_v, row_f in zip(ws_values.iter_rows(), ws_formulas.iter_rows()):
        for cell_v, cell_f in zip(row_v, row_f):
            formula = (
                cell_f.value
                if isinstance(cell_f.value, str) and str(cell_f.value).startswith("=")
                else None
            )
            value = cell_v.value
            colour = fill_colour(cell_v)
            if value is None and formula is None and colour is None:
                continue
            if len(lines) >= MAX_CELLS:
                truncated = True
                break
            parts = [f"{cell_v.coordinate}:"]
            if value is not None:
                text = " ".join(str(value).split())
                if len(text) > MAX_VALUE_CHARS:
                    clipped.append(f"{cell_v.coordinate} (value, {len(text)} chars)")
                    text = text[:MAX_VALUE_CHARS]
                parts.append(f'value="{text}"')
            if formula:
                if len(formula) > MAX_FORMULA_CHARS:
                    clipped.append(
                        f"{cell_v.coordinate} (formula, {len(formula)} chars)"
                    )
                    formula = formula[:MAX_FORMULA_CHARS]
                parts.append(f"formula={formula}")
            if colour:
                parts.append(f"fill={colour}")
            lines.append(" ".join(parts))
        if truncated:
            break

    warnings = []
    if truncated:
        warnings.append(
            f"TRUNCATED at {MAX_CELLS} cells - sheet is "
            f"{ws_values.max_row}x{ws_values.max_column}. The dropped cells are "
            "invisible to the model, so any interpretation describes PART of "
            "this sheet. Raise MAX_CELLS, or do not interpret it cell by cell."
        )
    if clipped:
        warnings.append(
            f"CLIPPED {len(clipped)} cell(s) at MAX_VALUE_CHARS/"
            f"MAX_FORMULA_CHARS: {', '.join(clipped[:10])}"
            f"{' ...' if len(clipped) > 10 else ''}. A clipped cell loses "
            "whatever came after the cut - check these before trusting any "
            "statement drawn from them."
        )

    merges = [str(r) for r in ws_formulas.merged_cells.ranges]
    header = [f"SHEET: {ws_values.title}"]
    if merges:
        header.append(f"MERGED RANGES: {', '.join(merges[:40])}")
    header.append(f"CELLS ({len(lines)}{'+, TRUNCATED' if truncated else ''}):")
    return "\n".join(header + lines), warnings


def interpret(client, model, payload):
    """One API call per sheet, schema-constrained.

    Running out of output budget is not an error the API raises - the call
    returns 200 with stop_reason='max_tokens' and JSON cut mid-structure. Left
    to json.loads that surfaces as 'Unterminated string at column 7823', which
    points at everything except the cause. Named here instead: the sheet had
    more to say than MAX_OUTPUT_TOKENS allowed.
    """
    response = client.messages.create(
        model=model,
        max_tokens=MAX_OUTPUT_TOKENS,
        system=SYSTEM,
        messages=[{"role": "user", "content": payload}],
        output_config={"format": {"type": "json_schema", "schema": SCHEMA}},
    )
    if response.stop_reason == "max_tokens":
        raise RuntimeError(
            f"the response hit MAX_OUTPUT_TOKENS ({MAX_OUTPUT_TOKENS}) after "
            f"{response.usage.output_tokens} tokens, so its JSON is cut off "
            "and nothing from this sheet can be used. The sheet has more to "
            "say than the budget allows: raise MAX_OUTPUT_TOKENS, or split "
            "the sheet and interpret it in parts."
        )
    text = next(b.text for b in response.content if b.type == "text")
    return json.loads(text)


def render_review(name, sheet_name, result):
    """Human-readable draft. Low confidence and uncertainties first: those are
    what your review time should be spent on."""
    out = [
        f"# {name} - {sheet_name}",
        "",
        f"**Kind:** {result['kind']}",
        "",
        result["sheet_purpose"],
        "",
        "> DRAFT. Every statement below was inferred by a model from cell "
        "positions. Check the cited cells before any of this is indexed.",
        "",
    ]

    if result["uncertain"]:
        out += ["## Could not interpret - answer these first", ""]
        for u in result["uncertain"]:
            out.append(f"- **{', '.join(u['cells'])}** - {u['question']}")
        out.append("")

    order = {"low": 0, "medium": 1, "high": 2}
    items = sorted(result["durable"], key=lambda d: order.get(d["confidence"], 0))
    low = [d for d in items if d["confidence"] != "high"]
    if low:
        out += ["## Durable knowledge - CHECK THESE (medium/low confidence)", ""]
        for d in low:
            out += _item(d)
    high = [d for d in items if d["confidence"] == "high"]
    if high:
        out += ["## Durable knowledge - high confidence", ""]
        for d in high:
            out += _item(d)

    if result["instance_values"]:
        out += [
            "## Instance values - EXCLUDE from the index",
            "",
            "One run of the sheet. These change whenever someone re-enters inputs.",
            "",
        ]
        for v in result["instance_values"]:
            out.append(f"- {v['label']} ({v['cell']}): {v['value']}")
        out.append("")
    return "\n".join(out) + "\n"


def _item(d):
    lines = [f"- **{d['statement']}**"]
    lines.append(f"  - category: {d['category']}, confidence: {d['confidence']}")
    lines.append(f"  - cells: {', '.join(d['source_cells'])} - {d['evidence']}")
    if d["formula"]:
        lines.append(f"  - formula: `{d['formula']}`")
    lines.append("")
    return lines


def main():
    parser = argparse.ArgumentParser(
        description="Interpret spreadsheets with openpyxl + Claude. Produces a draft to review."
    )
    parser.add_argument("--input", type=Path, default=Path("data/corpus"))
    parser.add_argument("--output", type=Path, default=Path("data/interp"))
    parser.add_argument("--model", default="claude-sonnet-5")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="write the payloads that would be sent and make no API call",
    )
    parser.add_argument(
        "--include-record-grids",
        action="store_true",
        help=(
            "interpret every sheet of a repeated form (all twelve monthly "
            "tabs). By default only the first of each template is interpreted "
            "and the rest are skipped."
        ),
    )
    args = parser.parse_args()

    # CWD first (running from the repo root), then next to the script. Neither
    # existing is fine: the variable may already be in the environment.
    load_dotenv()
    load_dotenv(Path(__file__).resolve().parent / ".env")

    if not args.input.is_dir():
        raise SystemExit(f"input directory not found: {args.input}")
    paths = sorted(p for p in args.input.iterdir() if p.suffix.lower() in SHEET_FORMATS)
    if not paths:
        raise SystemExit(f"no .xlsx/.xlsm in {args.input}")
    args.output.mkdir(parents=True, exist_ok=True)

    skipped = 0
    client = None
    if not args.dry_run:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            raise SystemExit(
                "ANTHROPIC_API_KEY is not set. Put it in a .env file at the "
                "repo root or next to this script, or set it in your shell. "
                "If you have a .env, check the key is named exactly "
                "ANTHROPIC_API_KEY and that python-dotenv is installed."
            )
        import anthropic

        client = anthropic.Anthropic()

    for path in paths:
        print(f"{path.name}", flush=True)
        try:
            wb_values = load_workbook(path, data_only=True)
            wb_formulas = load_workbook(path, data_only=False)
        except Exception as exc:
            print(f"  FAILED to open: {exc}", file=sys.stderr, flush=True)
            continue

        # Grouped up front: whether a sheet is skipped depends on whether an
        # earlier sheet in the same workbook already stood in for it.
        stands_in_for = {}
        if not args.include_record_grids:
            grids = {}
            for sheet_name in wb_values.sheetnames:
                shapes = record_grid_shapes(wb_formulas[sheet_name])
                if shapes:
                    grids[sheet_name] = shapes
            stands_in_for = group_record_grids(grids)

        for sheet_name in wb_values.sheetnames:
            stem = f"{path.stem}__{sheet_name}".replace(" ", "_").replace("/", "_")
            try:
                representative = stands_in_for.get(sheet_name)
                if representative and representative != sheet_name:
                    skipped += 1
                    print(
                        f"  {sheet_name}: SKIPPED - same template as "
                        f"'{representative}', which is being interpreted for "
                        "the whole group. Its own entries are instance data "
                        "and belong in a queryable table. Pass "
                        "--include-record-grids to interpret it too.",
                        flush=True,
                    )
                    continue
                if representative:
                    kin = [
                        n for n, r in stands_in_for.items() if r == sheet_name
                    ]
                    if len(kin) > 1:
                        print(
                            f"  {sheet_name}: record grid - interpreting as the "
                            f"representative of {len(kin)} sheets sharing this "
                            f"template ({', '.join(kin)})",
                            flush=True,
                        )

                payload, warnings = extract_sheet(
                    wb_values[sheet_name], wb_formulas[sheet_name]
                )
                for warning in warnings:
                    print(f"  {sheet_name}: {warning}", flush=True)
                (args.output / f"{stem}.payload.txt").write_text(
                    payload, encoding="utf-8"
                )
                if args.dry_run:
                    print(f"  {sheet_name}: payload written ({len(payload)} chars)")
                    continue

                result = interpret(client, args.model, payload)
                (args.output / f"{stem}.json").write_text(
                    json.dumps(result, indent=2), encoding="utf-8"
                )
                (args.output / f"{stem}.review.md").write_text(
                    render_review(path.name, sheet_name, result), encoding="utf-8"
                )
                flag = len([d for d in result["durable"] if d["confidence"] != "high"])
                print(
                    f"  {sheet_name}: {result['kind']}, "
                    f"{len(result['durable'])} durable ({flag} to check), "
                    f"{len(result['uncertain'])} uncertain",
                    flush=True,
                )
            except Exception as exc:
                traceback.print_exc()
                print(f"  {sheet_name}: FAILED: {exc}", file=sys.stderr, flush=True)

    if skipped:
        print(
            f"\n{skipped} sheet(s) skipped as duplicates of a template already "
            "interpreted (--include-record-grids to interpret them too)"
        )
    print(f"\n-> {args.output}   read the .review.md files before indexing anything")
    return 0


if __name__ == "__main__":
    sys.exit(main())
