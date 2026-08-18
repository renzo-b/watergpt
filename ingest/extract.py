"""Structural decomposition. No model call, no interpretation, no cost.

This is the half of the pipeline that a model has no business doing. Where the
tables and figures are, which page each sits on, where one section ends and the
next begins - docling and openpyxl already know, exactly, from the file itself.
Asking a model to infer it would be slower, dearer, and wrong in ways nothing
downstream could detect.

So the split down the middle of this pipeline is:

    extract.py    what parts exist, and where     - structural, free
    interpret.py  what each part means            - model, paid

PDFs and .docx come from the docling parse cached in the plant's parsed/
directory by inspect_parse.py, reusing the JSON you already reviewed rather than
reconverting - a 57-page scan costs minutes and the parse is the artefact you
checked. Anything without a cached parse is converted here with inspect_parse's
own converter settings, imported rather than copied so the two cannot drift.

Images are the one thing the cache does not hold. inspect_parse ran without
generate_picture_images, so every picture carries a page and a bounding box but
no pixels. Rather than reconvert the whole corpus to get them, the page region
is cropped from the source PDF on demand with pypdfium2 - which docling already
depends on, so it costs no new dependency and no re-parse.
"""

import base64
import hashlib
import io
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import plant as plant_paths  # noqa: E402 - needs ROOT on sys.path first

# Page furniture. Running headers and footers repeat on every page and say
# nothing about content; carrying them would inflate every section with the
# document title and a page number.
FURNITURE = {"page_header", "page_footer"}

# Rendering scale for cropped figures. 2x the PDF's nominal 72dpi is enough for
# a model to read axis labels on a pump curve without producing an image so
# large it dominates the request.
IMAGE_SCALE = 2.0

# Long edge of a figure after cropping. Rendering at 2x then shrinking beats
# rendering small: the resample averages whole pixels rather than sampling a
# thinly-rasterised curve, so hairlines and axis labels survive. Above roughly
# this width the API bills more tokens for detail a description will not use -
# a full-page figure at 2x is 600KB, and a document with 27 of them would ship
# 16MB of PNG to describe 27 diagrams.
MAX_IMAGE_PX = 1100


def file_hash(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()[:16]


def file_type(path):
    return Path(path).suffix.lower().lstrip(".")


class Part:
    """One extracted part, before any model has seen it.

    Mirrors schema.Component but carries the raw text and the image bytes,
    which are sent to the model and then dropped rather than stored.
    """

    def __init__(self, component_id, kind, locator, title="", text="",
                 page_start=None, page_end=None, image=None):
        self.component_id = component_id
        self.kind = kind
        self.locator = locator
        self.title = title
        self.text = text
        self.page_start = page_start
        self.page_end = page_end
        self.image = image  # PNG bytes, pictures only

    def __repr__(self):
        return (f"<Part {self.component_id} {self.kind} {self.locator!r} "
                f"{len(self.text)}ch>")


# --------------------------------------------------------------------------
# PDF and Word
# --------------------------------------------------------------------------

def cached_parse(path, plant=None):
    """The reviewed DoclingDocument for this file, or None."""
    candidate = plant_paths.parsed_dir(plant) / (Path(path).stem + ".json")
    if candidate.exists():
        return json.loads(candidate.read_text(encoding="utf-8"))
    return None


def convert_now(path):
    """Convert a file with no cached parse, using inspect_parse's settings."""
    from types import SimpleNamespace

    sys.path.insert(0, str(ROOT / "scripts"))
    import inspect_parse  # noqa: PLC0415 - late: it configures torch on import

    args = SimpleNamespace(no_ocr=False, table_mode="accurate", no_cell_matching=False)
    converter = inspect_parse.build_converter(args)
    result = converter.convert(str(path))
    return result.document.export_to_dict()


def page_of(item):
    prov = item.get("prov") or []
    return prov[0].get("page_no") if prov else None


def crop_picture(pdf_path, item):
    """Render the page region a picture occupies, as PNG bytes.

    The cached parse has the box but not the pixels. Cropping from the source
    beats reconverting the corpus with generate_picture_images: it is seconds
    rather than minutes and leaves the reviewed parse untouched.
    """
    import pypdfium2 as pdfium

    prov = (item.get("prov") or [None])[0]
    if not prov:
        return None
    box = prov["bbox"]

    pdf = pdfium.PdfDocument(str(pdf_path))
    try:
        page = pdf[prov["page_no"] - 1]
        height = page.get_height()
        image = page.render(scale=IMAGE_SCALE).to_pil()

        # docling boxes are BOTTOMLEFT-origin: t and b are measured up from the
        # foot of the page, so the top of the box is the larger number. PIL is
        # TOPLEFT-origin, hence the subtraction.
        top, bottom = max(box["t"], box["b"]), min(box["t"], box["b"])
        left, right = min(box["l"], box["r"]), max(box["l"], box["r"])
        crop = image.crop((
            int(left * IMAGE_SCALE),
            int((height - top) * IMAGE_SCALE),
            int(right * IMAGE_SCALE),
            int((height - bottom) * IMAGE_SCALE),
        ))
        if crop.width < 2 or crop.height < 2:
            return None

        longest = max(crop.width, crop.height)
        if longest > MAX_IMAGE_PX:
            ratio = MAX_IMAGE_PX / longest
            crop = crop.resize(
                (max(1, int(crop.width * ratio)), max(1, int(crop.height * ratio))),
                resample=1,  # PIL.Image.LANCZOS, without importing the enum
            )

        # JPEG at high quality. Most figures in this corpus are scans, where
        # PNG stores photographic noise losslessly and costs ten times the
        # bytes for no readable detail: a 777x592 crop is 479KB as PNG and
        # 40KB here. Quality is kept high because the detail that matters is
        # thin strokes and small axis labels on the drawings that are line art.
        buf = io.BytesIO()
        crop.convert("RGB").save(buf, format="JPEG", quality=90, optimize=True)
        return buf.getvalue()
    finally:
        pdf.close()


def caption_of(doc, item):
    """Caption text for a table or picture, resolved through its reference."""
    out = []
    for ref in item.get("captions", []) or []:
        parts = (ref.get("$ref") or "").lstrip("#/").split("/")
        if len(parts) == 2 and parts[0] in doc:
            try:
                out.append(doc[parts[0]][int(parts[1])].get("text", ""))
            except (ValueError, IndexError, KeyError, TypeError):
                pass
    return " ".join(t for t in out if t).strip()


def render_table(item):
    """A table as pipe-delimited rows, which is what a model reads best."""
    grid = (item.get("data") or {}).get("grid") or []
    rows = []
    for row in grid:
        cells = [str(c.get("text", "")).strip() for c in row]
        if any(cells):
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def extract_document(path, plant=None):
    """PDF or .docx -> ordered parts: sections, tables, figures, formulas."""
    path = Path(path)
    doc = cached_parse(path, plant)
    from_cache = doc is not None
    if doc is None:
        doc = convert_now(path)

    parts = []
    current = None
    counter = 0

    # Sections: contiguous prose under one heading. Text items are fine-grained
    # (the median is 22 characters - a list bullet, one line of a numbered
    # step), so a component per text item would be useless as a locator. A
    # heading and the body beneath it is the unit an operator asks for.
    for item in doc.get("texts", []):
        label = item.get("label")
        if label in FURNITURE:
            continue
        text = (item.get("text") or "").strip()
        if not text:
            continue
        page = page_of(item)

        if label == "formula":
            counter += 1
            parts.append(Part(
                component_id=f"formula.{counter}",
                kind="formula",
                locator=f"page {page}" if page else "",
                text=text,
                page_start=page, page_end=page,
            ))
            continue

        if label == "section_header" or current is None:
            counter += 1
            current = Part(
                component_id=f"section.{counter}",
                kind="section",
                locator=f"page {page}" if page else "",
                title=text if label == "section_header" else "",
                text="" if label == "section_header" else text,
                page_start=page, page_end=page,
            )
            parts.append(current)
        else:
            current.text = (current.text + "\n" + text).strip()
            if page:
                current.page_end = page

    # Section locators only now that each section's extent is known.
    for part in parts:
        if part.kind == "section" and part.page_start:
            part.locator = (
                f"page {part.page_start}"
                if part.page_end in (None, part.page_start)
                else f"pages {part.page_start}-{part.page_end}"
            )

    for i, item in enumerate(doc.get("tables", []), 1):
        page = page_of(item)
        parts.append(Part(
            component_id=f"table.{i}",
            kind="table",
            locator=f"page {page}" if page else "",
            title=caption_of(doc, item),
            text=render_table(item),
            page_start=page, page_end=page,
        ))

    for i, item in enumerate(doc.get("pictures", []), 1):
        page = page_of(item)
        image = None
        if path.suffix.lower() == ".pdf":
            try:
                image = crop_picture(path, item)
            except Exception:  # noqa: BLE001 - a figure that will not crop is
                image = None   # described from its caption, or not at all
        parts.append(Part(
            component_id=f"picture.{i}",
            kind="image",
            locator=f"page {page}" if page else "",
            title=caption_of(doc, item),
            page_start=page, page_end=page,
            image=image,
        ))

    parts.sort(key=lambda p: (p.page_start or 0, p.component_id))
    source_chars = sum(
        len(t.get("text") or "") for t in doc.get("texts", [])
        if t.get("label") not in FURNITURE
    )
    return parts, {"from_cache": from_cache, "source_chars": source_chars}


# --------------------------------------------------------------------------
# Spreadsheets and CSV
# --------------------------------------------------------------------------

def extract_workbook(path, plant=None):
    """One part per sheet, carrying the structural dump logs/convert.py sends.

    Whether a sheet is a log grid or a calculation sheet is left to the model:
    that is the one routing call in this pipeline that genuinely needs
    judgement, because the two look alike to any rule you would write - a dated
    column and a lot of numbers describes both a monthly record sheet and a
    design worksheet with a units table down the side.
    """
    from openpyxl import load_workbook

    from logs.convert import render_sheet

    wb = load_workbook(path, data_only=True)
    parts = []
    for i, name in enumerate(wb.sheetnames, 1):
        parts.append(Part(
            component_id=f"sheet.{i}",
            kind="table",  # provisional; the model may reclassify it as `log`
            locator=f"sheet {name!r}",
            title=name,
            text=render_sheet(wb[name]),
        ))
    source_chars = sum(
        len(str(c.value)) for name in wb.sheetnames
        for row in wb[name].iter_rows() for c in row if c.value is not None
    )
    return parts, {"from_cache": False, "source_chars": source_chars}


def extract_csv(path, plant=None):
    """One part: a CSV is a single table by definition."""
    import pandas as pd

    df = pd.read_csv(path)
    text = (
        f"COLUMNS ({len(df.columns)}): {', '.join(map(str, df.columns))}\n"
        f"ROWS: {len(df)}\n\nFIRST ROWS:\n{df.head(20).to_string()}"
    )
    return [Part(
        component_id="table.1",
        kind="table",
        locator=Path(path).name,
        title=Path(path).stem,
        text=text,
    )], {"from_cache": False, "source_chars": int(df.size)}


DISPATCH = {
    "pdf": extract_document,
    "docx": extract_document,
    "doc": extract_document,
    "xlsx": extract_workbook,
    "xlsm": extract_workbook,
    "csv": extract_csv,
}


def extract(path, plant=None):
    """Decompose any supported file into parts. Raises on an unknown type."""
    kind = file_type(path)
    if kind not in DISPATCH:
        raise ValueError(f"unsupported file type {kind!r}: {path}")
    return DISPATCH[kind](path, plant)


def encode_image(png_bytes):
    return base64.standard_b64encode(png_bytes).decode("ascii")
