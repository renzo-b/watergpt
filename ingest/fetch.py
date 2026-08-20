"""Reading what the catalogue points at. No model call, no exec.

The catalogue is an index: it says a thing exists and where. This is the other
half - given a locator the catalogue printed, return the source behind it.

    fetch("Rutledge Creek WWTP O&M Manual 2007-01.pdf", "pages 93-103")
    fetch("excel_1s_simple.xlsx", "sheet 'CT'")

Why this has to exist. A description can route a question and cannot answer
one. "Covers UV disinfection, pages 93-103" is enough to know where to look and
nothing like enough to say what the lamp cleaning interval is. Without a fetch
path the coarse tier is a card catalogue for a library with no books, and the
larger the document the more completely that is true.

Content is read from the same sources ingest/ wrote from - the cached docling
parse for a document, the workbook for a spreadsheet, parquet for a log - so
what comes back is the source, not a stored copy that can drift from it.
"""

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import plant as plant_paths  # noqa: E402
from ingest.extract import FURNITURE, caption_of, render_table  # noqa: E402

# Characters returned per fetch. Generous - the point of fetching is to read
# the thing - but bounded, because a chapter of a 660-page manual can be tens
# of thousands of tokens and the answer rarely needs all of it at once.
MAX_FETCH_CHARS = 40_000


def parse_locator(locator):
    """Turn a catalogue locator back into something addressable.

    Accepts what catalogue_entry prints: "page 7", "pages 93-103", "sheet 'CT'",
    and the log form "sheet 'Jan.' (2016-01-01 to 2016-01-31, 31 rows)" whose
    trailing facts are pandas-computed and not part of the address.
    """
    text = str(locator or "").strip().strip("[]")

    sheet = re.search(r"sheet\s+['\"](.+?)['\"]", text)
    if sheet:
        return ("sheet", sheet.group(1), None)

    span = re.search(r"pages?\s*\.?\s*(\d+)\s*[-–]\s*(\d+)", text)
    if span:
        return ("pages", int(span.group(1)), int(span.group(2)))

    one = re.search(r"pages?\s*\.?\s*(\d+)", text)
    if one:
        return ("pages", int(one.group(1)), int(one.group(1)))

    return (None, None, None)


def find_document(document, plant=None):
    """The manifest entry for a document, matched on filename."""
    from ingest.pipeline import read_manifest

    want = Path(str(document)).name.casefold()
    for entry in read_manifest(plant):
        if Path(entry.file_path).name.casefold() == want:
            return entry
    return None


def pages_text(entry, start, end, plant=None):
    """Source text for a page range, read from the cached docling parse."""
    cached = plant_paths.parsed_dir(plant) / (Path(entry.file_path).stem + ".json")
    if not cached.exists():
        return (
            f"No cached parse for {Path(entry.file_path).name}. It was ingested "
            "from a parse that has since been deleted; re-run ingest to rebuild it."
        )
    doc = json.loads(cached.read_text(encoding="utf-8"))

    def on_page(item):
        prov = item.get("prov") or []
        return prov and start <= prov[0].get("page_no", -1) <= end

    out, current = [], None
    for item in doc.get("texts", []):
        if item.get("label") in FURNITURE or not on_page(item):
            continue
        text = (item.get("text") or "").strip()
        if not text:
            continue
        page = item["prov"][0]["page_no"]
        if page != current:
            out.append(f"\n[page {page}]")
            current = page
        out.append(text)

    for item in doc.get("tables", []):
        if not on_page(item):
            continue
        page = item["prov"][0]["page_no"]
        caption = caption_of(doc, item)
        out.append(f"\n[page {page}] TABLE{': ' + caption if caption else ''}")
        out.append(render_table(item))

    if not out:
        return f"Nothing was extracted from pages {start}-{end}."
    return "\n".join(out)


def log_text(entry, sheet_name, columns=None, plant=None):
    """One converted log sheet: its per-column facts, or the rows of a column.

    NOT the whole dataframe. A month of this plant's daily log is 31 rows by
    133 columns - 238,000 characters rendered - so dumping it truncates after
    about four days and answers "the highest turbidity in January" with the
    first week of it. Worse, it truncates silently enough that only a careful
    reader notices the month is missing.

    So the default is the column index: every column with the unit, min, max
    and null fraction that pandas computed when the sheet was converted. That
    is what most questions about a log actually want - an extreme, a range,
    whether a column is even recorded - and those numbers are facts from the
    data rather than a model's reading of a wall of text.

    Ask for `columns` to get the rows themselves, narrowed to the columns whose
    names contain that text. A dozen columns of dailies fit easily; all 133
    never will.
    """
    from logs import find, load_log

    sheet = find(sheet_name, file_hash=entry.file_hash, plant=plant)
    if sheet is None:
        return f"No converted log sheet named {sheet_name!r}."

    if columns:
        try:
            df = load_log(sheet_name, file_hash=entry.file_hash, plant=plant)
        except LookupError as exc:
            return str(exc)
        wanted = [c for c in df.columns if columns.casefold() in str(c).casefold()]
        if not wanted:
            return (
                f"No column in {sheet_name!r} contains {columns!r}. Fetch "
                "without `columns` to see the column index."
            )
        narrowed = df[wanted]
        return (
            f"Converted log sheet {sheet_name!r}, {len(narrowed)} rows x "
            f"{len(wanted)} column(s) matching {columns!r}:\n\n"
            f"{narrowed.to_string()}"
        )

    lines = [
        f"Converted log sheet {sheet_name!r}: {sheet.row_count} rows x "
        f"{len(sheet.columns)} columns.",
    ]
    if sheet.coverage_start and sheet.coverage_end:
        lines.append(
            f"Covers {sheet.coverage_start:%Y-%m-%d} to "
            f"{sheet.coverage_end:%Y-%m-%d}"
            + (f", {sheet.sampling_interval}." if sheet.sampling_interval else ".")
        )
    lines += [
        "",
        "Column index. min/max/blank were computed by pandas from the data, "
        "not estimated. To see the daily rows of a column, fetch this sheet "
        "again with `columns` set to part of its name.",
        "",
        f"{'column':<74} {'unit':<8} {'min':>10} {'max':>10}  blank",
    ]
    for col in sheet.columns:
        unit = col.unit or ""
        lo = "" if col.min is None else f"{col.min:.4g}"
        hi = "" if col.max is None else f"{col.max:.4g}"
        lines.append(
            f"{str(col.name)[:74]:<74} {unit[:8]:<8} {lo:>10} {hi:>10}  "
            f"{col.null_frac:.0%}"
        )
    return "\n".join(lines)


def sheet_text(entry, sheet_name, columns=None, plant=None):
    """Source for one spreadsheet sheet: log facts, or cells with formulas."""
    component = next(
        (
            c for c in entry.components
            if c.title == sheet_name or sheet_name.casefold() in c.locator.casefold()
        ),
        None,
    )

    if component is not None and component.kind == "log":
        return log_text(entry, sheet_name, columns, plant)

    from openpyxl import load_workbook

    from ingest.extract import render_sheet

    values = load_workbook(entry.file_path, data_only=True)
    formulas = load_workbook(entry.file_path, data_only=False)
    if sheet_name not in values.sheetnames:
        return (
            f"No sheet named {sheet_name!r}. This workbook has: "
            f"{', '.join(values.sheetnames)}"
        )
    return render_sheet(values[sheet_name], formulas[sheet_name])


def fetch(document, locator, columns=None, plant=None):
    """Return the source behind one catalogue locator, as text."""
    entry = find_document(document, plant)
    if entry is None:
        from ingest.pipeline import read_manifest

        known = sorted(Path(e.file_path).name for e in read_manifest(plant))
        return (
            f"No ingested document named {document!r}. Available: "
            + (", ".join(known) if known else "(none)")
        )

    kind, a, b = parse_locator(locator)
    if kind == "pages":
        body = pages_text(entry, a, b, plant)
    elif kind == "sheet":
        body = sheet_text(entry, a, columns, plant)
    else:
        return (
            f"Could not read {locator!r} as a location. Use a locator exactly as "
            "the catalogue prints it: 'page 7', 'pages 93-103', or \"sheet 'CT'\"."
        )

    if len(body) > MAX_FETCH_CHARS:
        body = (
            body[:MAX_FETCH_CHARS]
            + f"\n\n... truncated at {MAX_FETCH_CHARS:,} characters. Fetch a "
            "narrower page range to see the rest."
        )
    return f"{Path(entry.file_path).name} @ {locator}\n\n{body}"
