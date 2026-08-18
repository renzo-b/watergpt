"""One LLM call per workbook. It writes a converter; code runs it to parquet.

    python -m logs.convert --input documents/retrieval_test
    python -m logs.convert --input book.xlsx --dry-run   # payload only, no call

The model sees a structural dump of every sheet in one workbook and returns a
single converter plus a description per sheet. Code then runs the converter,
validates each frame, writes parquet, and appends manifest rows.

Two things the model is never asked to do:

  - emit cell values. It writes code; code reads the sheet. Transcribing
    thousands of cells would be slower, dearer, and would put transcription
    errors into the source of truth.
  - state a statistic. Row counts, date ranges, minima and distinct values are
    computed by pandas afterwards. A model asked for them will approximate,
    and an approximated fact is worse than a missing one because nothing
    downstream can tell.

Failure is per sheet and quiet: a frame that fails validation is recorded
unconverted with the error and left out of the catalogue. One retry with the
failure text appended, then move on. Converting fourteen of fifteen sheets is
a good outcome; taking down the run because of the fifteenth is not.
"""

import argparse
import hashlib
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

import plant as plant_paths
from logs.loader import (
    append_manifest,
    describe_frame,
    parquet_path,
    read_manifest,
)
from logs.schema import ColumnInfo, SheetEntry, validate_frame

try:
    from dotenv import load_dotenv
except ImportError:  # optional - the shell environment works just as well
    def load_dotenv(*_args, **_kwargs):
        return False

DEFAULT_MODEL = "claude-opus-5"

# One converter covering every shape in a workbook runs long, and the
# per-sheet descriptions ride in the same response. A response cut here cannot
# be parsed or partly recovered, so the ceiling is generous and the run says so
# loudly if it still trips. 32000 was not enough for a 15-sheet workbook.
MAX_OUTPUT_TOKENS = 64000

# How much of each sheet the model sees. The header region has to be complete,
# since that is what the converter is written against; the data region only
# needs enough rows to show the shape, and the footer enough to show where the
# table stops.
HEADER_ROWS = 10
DATA_SAMPLE_ROWS = 3
FOOTER_ROWS = 4
MAX_CELLS_PER_SHEET = 900
MAX_MERGES_PER_SHEET = 120

SYSTEM = """\
You write one Python module that converts every sheet of an Excel workbook \
into tidy pandas DataFrames.

You will be given a structural dump of each sheet: dimensions, merged ranges, \
and cell values by coordinate for the header region, a few data rows and the \
last rows. Sheets in one workbook often share a template; write one converter \
that handles every shape present.

Define exactly one public function:

    convert(path) -> dict[str, pandas.DataFrame]

Keyed by sheet name, one entry per sheet you can convert. Omit a sheet you \
cannot make sense of rather than returning something you do not believe — an \
omitted sheet is recorded as unconverted, which is a fine outcome.

Each DataFrame: one row per observation, one column per measurement. If the \
sheet has an observation date, make it a DatetimeIndex named "date"; if it has \
none, leave a default index.

Rules that matter more than they look:

- Expand merged ranges before reading any header. A merged cell holds its \
value only at the anchor; every other covered cell reads as None.
- Column names must be unique and readable. Where a leaf label repeats under \
different groups (three columns all labelled "Meter Reading"), join the levels \
above it. Join with " | ".
- Do not hardcode row counts or a number of footer rows to skip. Detect where \
data ends, so the same code works for a 28-, 30- and 31-row month.
- If the observation date is not stored as a date, reconstruct it. It may be \
split across a period label in a title cell and a day number in a column. If a \
sheet repeats the same day number in several columns, prefer the value most of \
them agree on — a single column can carry a typo.
- Use openpyxl. pd.read_excel cannot express merged multi-row headers.
- Coerce numeric columns, but only where coercion does not destroy the column: \
a column of operator initials must stay text.
- Write the module self-contained: imports at the top, no code outside \
functions, no printing, no writing files, no network.

Describe each sheet for an operator who is trying to find it, in the words \
they would use. State no statistic about any sheet - no row count, date range, \
distinct values, minimum or maximum. Those are computed from your converter \
afterwards."""

SCHEMA = {
    "type": "object",
    "properties": {
        "converter_source": {
            "type": "string",
            "description": (
                "The complete Python module defining convert(path). "
                "Plain source, no markdown fences."
            ),
        },
        "converter_notes": {
            "type": "string",
            "description": (
                "Anything about the workbook that forced a decision, any sheet "
                "you chose to omit and why, and anything you were unsure of."
            ),
        },
        "sheets": {
            "type": "array",
            "description": "One entry per sheet you converted.",
            "items": {
                "type": "object",
                "properties": {
                    "sheet_name": {"type": "string"},
                    "description": {
                        "type": "string",
                        "description": (
                            "One paragraph: what this sheet records, in plant "
                            "terms, using the words an operator would search by."
                        ),
                    },
                    "column_gloss": {
                        "type": "array",
                        "description": (
                            "Plain-English meaning per column or column group. "
                            "Cover the groups and any abbreviation. Do not "
                            "enumerate every leaf of a repeating group."
                        ),
                        "items": {
                            "type": "object",
                            "properties": {
                                "column": {"type": "string"},
                                "meaning": {"type": "string"},
                            },
                            "required": ["column", "meaning"],
                            "additionalProperties": False,
                        },
                    },
                },
                "required": ["sheet_name", "description", "column_gloss"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["converter_source", "converter_notes", "sheets"],
    "additionalProperties": False,
}


def file_hash(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()[:16]


def populated_cells(ws):
    return sum(1 for row in ws.iter_rows() for c in row if c.value is not None)


def render_sheet(ws):
    """Structural dump of one sheet, scoped to what writing a converter needs."""
    last = ws.max_row
    wanted = list(range(1, HEADER_ROWS + 1))
    wanted += list(range(HEADER_ROWS + 1, min(HEADER_ROWS + 1 + DATA_SAMPLE_ROWS, last + 1)))
    wanted += [r for r in range(max(1, last - FOOTER_ROWS + 1), last + 1)]

    lines, truncated = [], False
    for r in sorted(set(wanted)):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            if len(lines) >= MAX_CELLS_PER_SHEET:
                truncated = True
                break
            value = cell.value
            if isinstance(value, str) and len(value) > 200:
                value = value[:200] + "..."
            lines.append(f"{cell.coordinate}: {value!r}")
        if truncated:
            break

    merges = [str(r) for r in ws.merged_cells.ranges]
    head = [
        f"--- SHEET: {ws.title!r} ---",
        f"DIMENSIONS: {ws.max_row} rows x {ws.max_column} columns",
        "MERGED RANGES: "
        + (", ".join(merges[:MAX_MERGES_PER_SHEET]) or "none")
        + (f" ... (+{len(merges) - MAX_MERGES_PER_SHEET} more)"
           if len(merges) > MAX_MERGES_PER_SHEET else ""),
        f"CELLS{' (TRUNCATED)' if truncated else ''}:",
    ]
    gap = sorted(set(range(1, last + 1)) - set(wanted))
    if gap:
        head.insert(
            3,
            f"NOTE: rows {gap[0]}-{gap[-1]} are omitted here and continue the "
            "pattern of the sampled data rows.",
        )
    return "\n".join(head + lines)


def render_workbook_payload(path):
    wb = load_workbook(path, data_only=True)
    parts = [
        f"WORKBOOK: {Path(path).name}",
        f"SHEETS ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}",
        "",
    ]
    parts += [render_sheet(wb[name]) for name in wb.sheetnames]
    return "\n\n".join(parts)


def call_model(client, model, payload, retry_note=None):
    content = payload
    if retry_note:
        content = (
            f"{payload}\n\nYOUR PREVIOUS CONVERTER WAS REJECTED. It ran against "
            f"this workbook and these sheets failed validation:\n\n{retry_note}\n\n"
            "Write a corrected converter. Fix the cause rather than the symptom, "
            "and do not suppress errors to make a sheet appear to convert."
        )

    with client.messages.stream(
        model=model,
        max_tokens=MAX_OUTPUT_TOKENS,
        system=SYSTEM,
        messages=[{"role": "user", "content": content}],
        output_config={"format": {"type": "json_schema", "schema": SCHEMA}},
    ) as stream:
        response = stream.get_final_message()

    if response.stop_reason == "max_tokens":
        raise RuntimeError(
            f"the response hit MAX_OUTPUT_TOKENS ({MAX_OUTPUT_TOKENS}) after "
            f"{response.usage.output_tokens} tokens, so its JSON is cut off and "
            "the converter cannot be recovered. Raise MAX_OUTPUT_TOKENS."
        )
    text = next(b.text for b in response.content if b.type == "text")
    return json.loads(text), response.usage


def run_converter(source, path):
    """Execute the converter once, here, at conversion time. Never at query time."""
    ns = {}
    exec(source, ns)  # noqa: S102 - model-authored converter, run once and inspected
    frames = ns["convert"](str(path))
    if not isinstance(frames, dict):
        raise TypeError(f"convert() returned {type(frames).__name__}, expected dict")
    return frames


def convert_workbook(path, client, model=DEFAULT_MODEL, log=print, plant=None):
    """One call, one converter, one parquet + manifest row per sheet."""
    path = Path(path)
    digest = file_hash(path)
    payload = render_workbook_payload(path)
    wb = load_workbook(path, data_only=True)
    populated = {name: populated_cells(wb[name]) for name in wb.sheetnames}

    result = frames = None
    retry_note = None
    for attempt in (1, 2):
        result, usage = call_model(client, model, payload, retry_note)
        log(f"  attempt {attempt}: {usage.input_tokens} in / {usage.output_tokens} out")
        try:
            frames = run_converter(result["converter_source"], path)
        except Exception as exc:  # noqa: BLE001 - handed back to the model verbatim
            retry_note = f"convert() raised {type(exc).__name__}: {exc}"
            log(f"  attempt {attempt} rejected: {retry_note}")
            frames = None
            continue

        bad = []
        for sheet, df in frames.items():
            failures, _ = validate_frame(df)
            if failures:
                bad.append(f"{sheet!r}: {'; '.join(failures)}")
        if not bad:
            break
        retry_note = "\n".join(bad)
        log(f"  attempt {attempt}: {len(bad)} sheet(s) failed validation")

    described = {s["sheet_name"]: s for s in (result or {}).get("sheets", [])}
    entries = []
    plant_paths.parquet_dir(plant).mkdir(parents=True, exist_ok=True)

    for sheet in wb.sheetnames:
        meta = described.get(sheet, {})
        gloss = {g["column"]: g["meaning"] for g in meta.get("column_gloss", [])}
        common = dict(
            file_path=str(path),
            file_hash=digest,
            sheet_name=sheet,
            generated_at=datetime.now(timezone.utc),
            description=meta.get("description", ""),
            column_gloss=gloss,
            converter_source=(result or {}).get("converter_source", ""),
        )

        df = (frames or {}).get(sheet)
        if df is None:
            entries.append(
                SheetEntry(
                    status="unconverted",
                    error="the converter did not return this sheet",
                    **common,
                )
            )
            continue

        failures, warnings = validate_frame(df)
        if failures:
            entries.append(
                SheetEntry(status="unconverted", error="; ".join(failures), **common)
            )
            continue

        out = parquet_path(digest, sheet, plant)
        df.to_parquet(out)
        facts = describe_frame(df)
        loaded = int(df.notna().sum().sum())
        entries.append(
            SheetEntry(
                status="converted",
                # Relative to the repo root: an absolute path bakes one
                # machine's home directory into a file that outlives it.
                parquet_path=plant_paths.relative(out),
                warnings=warnings,
                cell_coverage=(
                    round(loaded / populated[sheet], 3) if populated[sheet] else None
                ),
                row_count=facts["row_count"],
                columns=[ColumnInfo(**c) for c in facts["columns"]],
                coverage_start=facts["coverage_start"],
                coverage_end=facts["coverage_end"],
                sampling_interval=facts["sampling_interval"],
                distinct_values=facts["distinct_values"],
                **common,
            )
        )

    for entry in entries:
        append_manifest(entry, plant)
    return entries


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, help=".xlsx file or directory")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--plant", default=plant_paths.DEFAULT_PLANT,
                        help="which plant's data directory to write into")
    parser.add_argument("--force", action="store_true",
                        help="reconvert workbooks already in the manifest")
    parser.add_argument("--dry-run", action="store_true",
                        help="write the payload that would be sent and call nothing")
    args = parser.parse_args(argv)

    src = Path(args.input)
    books = sorted(src.glob("*.xlsx")) if src.is_dir() else [src]
    if not books:
        raise SystemExit(f"no .xlsx under {src}")

    if args.dry_run:
        for book in books:
            out = plant_paths.scratch_dir() / f"{book.stem}.payload.txt"
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_text(render_workbook_payload(book), encoding="utf-8")
            print(f"{out} ({out.stat().st_size:,} bytes)")
        return 0

    load_dotenv()
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise SystemExit("ANTHROPIC_API_KEY is not set")

    import anthropic

    client = anthropic.Anthropic()
    done = {e.file_hash for e in read_manifest(args.plant)}
    for book in books:
        if not args.force and file_hash(book) in done:
            print(f"{book.name}: already converted (use --force)")
            continue
        print(book.name)
        entries = convert_workbook(book, client, args.model, plant=args.plant)
        ok = [e for e in entries if e.status == "converted"]
        print(f"  {len(ok)}/{len(entries)} sheets converted")
        for e in entries:
            if e.status == "converted":
                print(f"    {e.sheet_name:<30} {e.row_count:>4} rows "
                      f"{len(e.columns):>4} cols  coverage={e.cell_coverage}")
            else:
                print(f"    {e.sheet_name:<30} UNCONVERTED: {e.error}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
